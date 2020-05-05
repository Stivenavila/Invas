import random
import unittest
from datetime import datetime, time
from fileinput import filename

from unipath import Path
from pyunitreport import HTMLTestRunner
from time import sleep, strftime, gmtime
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys


class PythonOrgSearch(unittest.TestCase):

    def setUp(self):
        self.driver = webdriver.Chrome("Driver_Data\chromedriver.exe")
        self.driver.implicitly_wait(60)
        self.driver.maximize_window()

    # @staticmethod
    # def highlight(element, effect_time, color, border):
    #   driver = str(element._parent)

    #  def apply_style(s):
    #     driver.execute_script("arguments[0].setAttribute('style', arguments[1]);",
    #                          element, s)

    # original_style = element.get_attribute('style')
    # apply_style("border: {0}px solid {1};".format(border, color))
    # time.sleep(effect_time)
    # apply_style(original_style)

    def test_search_in_python_org(self):

        f = Path('Driver_Data\Data.xlsx')
        try:
            f
        except FileNotFoundError:
            print("opss!!! el archivo no existe")

        wb = load_workbook(f)
        ws = wb.active
        driver = self.driver
        # obtain URL
        try:
            driver.get(ws['A2'].value)
        except Exception as e:
            print("URL: " + ws['A2'].value + "no es valida o no funciona" + str(e))

        # obtain user
        try:
            driver.find_element_by_name('name').send_keys(ws['B2'].value)
            driver.find_element_by_name('password').send_keys(ws['C2'].value)
        except Exception as e:
            print("User's not found" + str(e))

        # test_Login(self)
        filename = (strftime("%Y-%m-%d %H%M%S", gmtime()))
        sleep(0.5)
        driver.get_screenshot_as_file("Driver_Data\screanshot/1_login_" + filename + ".png")
        driver.find_element(By.XPATH, '//button[@class="btn btn-danger pull-right"]').click()
        driver.find_element(By.XPATH, '//input[@class="ui-grid-filter-input ui-grid-filter-input-0"]').click()
        driver.find_element(By.XPATH, '//input[@class="ui-grid-filter-input ui-grid-filter-input-0"]').send_keys(
            ws['D2'].value)
        driver.find_element(By.XPATH,
                            '//div[@class ="ui-grid-selection-row-header-buttons ui-grid-icon-ok ng-scope"]').click()
        driver.get_screenshot_as_file("Driver_Data\screanshot/2_select_site_" + filename + ".png")
        driver.find_element(By.XPATH, '//button[@class="btn btn-danger"]').click()

        # test_WMS_Management(self):
        driver.find_element(By.XPATH, "//a[contains(text(),'* GESTIÓN WMS')]").click()
        driver.get_screenshot_as_file("Driver_Data\screanshot/3_Management_wms_" + filename + ".png")

        # test_Picking_Management(self):
        driver.find_element(By.XPATH, '//*[@id="appRf"]/div[2]/div[2]/div/div/ul/li[13]').click()
        driver.get_screenshot_as_file("Driver_Data\screanshot/4_Management_piking_" + filename + ".png")

        # test_run_express_continuous_task(self):
        driver.find_element(By.XPATH, '//*[@id="appRf"]/div[2]/div[2]/div/div/ul/li[6]').click()
        # texto = text1.get_attribute('value')
        driver.get_screenshot_as_file("Driver_Data\screanshot/5_Continuous_task_" + filename + ".png")

        # alert_error
        try:
            # test_select_task(self):
            driver.find_element(By.XPATH, '(//input[@type="text"])[7]').click()
            driver.find_element(By.XPATH, '(//input[@type="text"])[7]').send_keys(ws['F2'].value, Keys.ENTER)
            # self.highlight('(//input[@type="text"])[7]', 3, "red", 5)
            driver.get_screenshot_as_file("Driver_Data\screanshot/6_Select_task_" + filename + ".png")
            driver.find_element(By.XPATH,
                                '//div[@class="ui-grid-selection-row-header-buttons ui-grid-icon-ok ng-scope"]').click()
            driver.find_element(By.XPATH, '//center//button[2]').click()
        except:
            popup = driver.find_element(By.XPATH, '//body[@id="appRf"]/div[2]/div[3]/div/div/div/div/b')
            popup2 = popup.text
            sleep(0.5)
            driver.get_screenshot_as_file("Driver_Data\screanshot/00_Alert_error" + filename + ".png")
            print("Opss!!, " + str(popup2))
            driver.close()

        # test_ForOrder(self):
        a = 1
        b = 7
        for i in range(ws['E2'].value):
            # test_Order_Piked(self):
            try:
                sleep(0.5)
                driver.get_screenshot_as_file("Driver_Data\screanshot/_Order's_" + filename + ".png")
                valorCant = driver.find_element(By.XPATH, '//*[@id="appRf"]/div[2]/div[2]/div/div/div[10]/input')
                cant = valorCant.get_attribute('value')
                sleep(0.8)
                driver.find_element(By.XPATH, '//*[@id="appRf"]/div[2]/div[2]/div/div/div[16]/input').send_keys(cant,
                                                                                                                Keys.ENTER)
            except:
                driver.get_screenshot_as_file("Driver_Data\screanshot/00_Error_Alert" + str(a) + filename + ".png")
                print('Error occurred: ' + 'Cantidad minima error')
                driver.close()

            # test_lPN_Rq(self):
            try:
                valor = 'LPNOUT' + str(random.randint(0, 10000))
                print(valor)
                sleep(2)
                driver.find_element(By.XPATH, '//*[@id="appRf"]/div[4]/div/div/div/div[2]/div[2]/input').send_keys(
                    valor)
                sleep(1)
                driver.get_screenshot_as_file(
                    "Driver_Data/screanshot/" + str(b) + "LpnOut" + str(a) + filename + ".png")
                driver.find_element(By.XPATH, '//div[@class="modal-footer"]//button[1]').click()
            except IOError:
                print('Error occurred : ' 'LPNOUT no fue escrito.')
                driver.close()
            sleep(2)
            # test_popUp_correct
            sleep(0.5)
            popup = driver.find_element(By.XPATH, '//body[@id="appRf"]/div[2]/div[3]/div/div/div/div/b')
            popup2 = popup.text
            print("Yeah!!, " + str(popup2))
            sleep(2)

        # test_Order_out(self):
        driver.find_element(By.XPATH, '//*[@id="appRf"]/div[2]/div[2]/div/div/center/button[1]').click()
        sleep(2)
        driver.get_screenshot_as_file("Driver_Data\screanshot/" + str(b) + "" + filename + ".png")
        driver.find_element(By.XPATH, '//*[@id="user-header"]/a/span').click()
        driver.find_element(By.XPATH, '//*[@id="user-header"]/ul/li/a/span').click()
        driver.get_screenshot_as_file("Driver_Data\screanshot/" + str(b) + "Close_" + filename + ".png")

    def tearDown(self):
        self.driver.close()


if __name__ == "__main__":
    unittest.main(testRunner=HTMLTestRunner(output=''))
